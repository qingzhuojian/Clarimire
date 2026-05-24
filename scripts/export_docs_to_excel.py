# -*- coding: utf-8 -*-
"""导出数据表与后端接口文档为 Excel"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "系统数据表与接口文档.xlsx"

TABLES = [
    ("users", "用户表", [
        ("id", "用户ID", "主键，自增"),
        ("username", "用户名", "唯一，登录账号"),
        ("password", "密码", "MD5 加密存储"),
        ("real_name", "真实姓名", "用户姓名"),
        ("role", "角色", "admin / inspector / public"),
        ("phone", "手机号", "联系电话"),
        ("status", "状态", "1 启用，0 禁用"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("reservoirs", "水库基础信息表", [
        ("id", "水库ID", "主键，自增"),
        ("reservoir_name", "水库名称", "水库名称，唯一索引"),
        ("location", "位置", "地理位置描述"),
        ("latitude", "纬度", "地图坐标纬度"),
        ("longitude", "经度", "地图坐标经度"),
        ("capacity", "库容", "水库总库容"),
        ("status", "状态", "normal / maintenance / risk"),
        ("construction_date", "建设日期", "水库建成日期"),
        ("last_maintenance_date", "最后维护日期", "最近一次维护日期"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("patrol_records", "巡查记录表", [
        ("id", "记录ID", "主键，自增"),
        ("date", "巡查日期", "签到/巡查日期"),
        ("time", "巡查时间", "签到/巡查时间"),
        ("reservoir_name", "水库名称", "关联水库"),
        ("latitude", "纬度", "签到位置纬度"),
        ("longitude", "经度", "签到位置经度"),
        ("address", "详细地址", "签到地址"),
        ("inspector", "巡查员", "巡查员姓名"),
        ("inspector_username", "巡查员工号", "巡查员账号"),
        ("status", "状态", "pending / processing / completed"),
        ("has_issue", "是否有问题", "布尔值"),
        ("issue_type", "问题类型", "问题分类"),
        ("issue_severity", "问题严重性", "low / medium / high / critical"),
        ("description", "问题描述", "问题详细说明"),
        ("has_photo", "是否有照片", "布尔值"),
        ("photo_urls", "照片URL", "JSON 数组"),
        ("reporter_name", "上报人", "上报人姓名"),
        ("reporter_role", "上报人角色", "admin / inspector / public"),
        ("assigned_inspector", "指派巡查员", "被指派处理人"),
        ("assignment_note", "指派备注", "任务派发说明"),
        ("assignment_time", "指派时间", "任务派发时间"),
        ("situation_description", "处理情况", "处理过程描述"),
        ("processing_result", "处理结果", "resolved / improved / ongoing / reported"),
        ("completion_time", "完成时间", "任务完成时间"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("warning_records", "预警记录表", [
        ("id", "预警ID", "主键，自增"),
        ("warning_type", "预警类型", "如 COD 超标、水位超限等"),
        ("warning_level", "预警级别", "low / medium / high / critical"),
        ("reservoir_id", "水库ID", "外键，关联 reservoirs"),
        ("reservoir_name", "水库名称", "水库名称"),
        ("latitude", "纬度", "预警位置纬度"),
        ("longitude", "经度", "预警位置经度"),
        ("description", "预警描述", "预警详细说明"),
        ("indicator_value", "指标值", "当前监测指标值"),
        ("threshold_value", "阈值", "触发预警的阈值"),
        ("status", "状态", "pending / processed / dismissed"),
        ("created_at", "创建时间", "预警生成时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("warning_thresholds", "预警阈值配置表", [
        ("id", "配置ID", "主键，自增"),
        ("cod_threshold", "COD阈值", "化学需氧量阈值，单位 mg/L"),
        ("ammonia_nitrogen_threshold", "氨氮阈值", "氨氮阈值，单位 mg/L"),
        ("total_phosphorus_threshold", "总磷阈值", "总磷阈值，单位 mg/L"),
        ("total_nitrogen_threshold", "总氮阈值", "总氮阈值，单位 mg/L"),
        ("permanganate_threshold", "高锰酸盐指数阈值", "高锰酸盐指数阈值，单位 mg/L"),
        ("flood_limit_water_level", "汛限水位", "汛限水位，单位米"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("layer_configs", "图层配置表", [
        ("id", "配置ID", "主键，自增"),
        ("layer_type", "图层类型标识", "唯一，如 reservoirs、rivers"),
        ("layer_name", "显示名称", "地图图层显示名"),
        ("color", "颜色", "图层颜色，如 #3388FF"),
        ("opacity", "透明度", "0–1 之间"),
        ("visible", "是否显示", "布尔值"),
        ("icon", "图标样式", "polygon / line / point 等"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("roles", "角色表", [
        ("id", "角色ID", "主键，自增"),
        ("role_name", "角色名称", "如管理员、巡查员"),
        ("role_key", "角色标识", "唯一，如 admin、inspector"),
        ("description", "角色说明", "角色描述"),
        ("permissions", "权限列表", "JSON 格式权限数组"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("simulation_configs", "模拟参数预设表", [
        ("id", "配置ID", "主键，自增"),
        ("config_name", "配置名称", "模拟方案名称"),
        ("diffusion_radius", "扩散半径", "单位米"),
        ("decay_coefficient", "衰减系数", "污染物衰减系数"),
        ("wind_speed", "风速", "单位 m/s"),
        ("wind_direction", "风向", "单位度"),
        ("water_flow_rate", "水流流速", "单位 m/s"),
        ("simulation_duration", "模拟时长", "单位秒"),
        ("is_default", "是否默认配置", "布尔值"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("annotations", "标注管理表", [
        ("id", "标注ID", "主键，自增"),
        ("user_id", "用户ID", "外键，关联 users"),
        ("title", "标注标题", "标注名称"),
        ("content", "标注内容", "标注详细内容"),
        ("latitude", "纬度", "标注位置纬度"),
        ("longitude", "经度", "标注位置经度"),
        ("color", "颜色", "标注颜色"),
        ("icon", "图标", "标注图标样式"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("water_situation", "水情管理表", [
        ("id", "ID", "主键，自增"),
        ("reservoir_name", "库名", "水库名称"),
        ("date", "日期", "水情记录日期时间"),
        ("water_level", "库水位", "单位米"),
        ("storage", "蓄水量", "单位万立方米"),
        ("avg_inflow", "日平均入库流量", "单位立方米/秒"),
        ("avg_outflow", "日平均出库流量", "单位立方米/秒"),
        ("yoy_increase", "比去年同期增减", "单位万立方米"),
        ("total_capacity", "总库容", "单位万立方米"),
        ("flood_level", "汛限水位", "单位米"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("section_monitor", "监测断面管理表", [
        ("id", "ID", "主键，自增"),
        ("monitor_point_name", "监测点名称", "断面监测点名称"),
        ("reservoir_name", "水库名称", "所属水库"),
        ("year", "年份", "监测年份"),
        ("month", "月份", "监测月份"),
        ("ammonia_nitrogen", "氨氮", "单位 mg/L"),
        ("permanganate_index", "高锰酸盐指数", "单位 mg/L"),
        ("cod", "化学需氧量", "单位 mg/L"),
        ("flow_rate", "流量", "单位立方米/秒"),
        ("water_depth", "水深", "单位米"),
        ("total_nitrogen", "总氮", "单位 mg/L"),
        ("total_phosphorus", "总磷", "单位 mg/L"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("inspection_tasks", "巡查任务表", [
        ("id", "任务ID", "主键，自增"),
        ("title", "任务标题", "任务名称"),
        ("description", "任务描述", "任务详细说明"),
        ("reservoir_name", "水库名称", "关联水库"),
        ("latitude", "纬度", "任务位置纬度"),
        ("longitude", "经度", "任务位置经度"),
        ("status", "状态", "pending / processing / completed"),
        ("creator_id", "创建人ID", "任务创建者"),
        ("creator_name", "创建人姓名", "创建者姓名"),
        ("assignee_id", "指派人ID", "执行人 ID"),
        ("assignee_name", "任务执行人", "执行人姓名"),
        ("deadline", "任务截止时间", "任务截止日期"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
    ("task_feedbacks", "任务反馈表", [
        ("id", "反馈ID", "主键，自增"),
        ("task_id", "任务ID", "外键，关联 inspection_tasks"),
        ("content", "反馈内容", "反馈文字说明"),
        ("photos", "照片URL", "JSON 数组"),
        ("inspector", "反馈人", "反馈人姓名"),
        ("inspector_username", "反馈人账号", "反馈人登录账号"),
        ("created_at", "创建时间", "反馈提交时间"),
    ]),
    ("issue_reports", "问题上报表", [
        ("id", "上报ID", "主键，自增"),
        ("reservoir_name", "水库名称", "关联水库"),
        ("description", "问题描述", "问题详细说明"),
        ("severity", "严重性", "low / medium / high / critical"),
        ("notes", "备注", "补充说明"),
        ("photos", "照片URL", "JSON 数组"),
        ("latitude", "纬度", "问题位置纬度"),
        ("longitude", "经度", "问题位置经度"),
        ("address", "详细地址", "问题位置地址"),
        ("reporter_name", "上报人", "上报人姓名"),
        ("reporter_role", "上报人角色", "admin / inspector / public"),
        ("reporter_username", "上报人账号", "上报人登录账号"),
        ("status", "状态", "pending / processing / completed"),
        ("assigned_inspector", "指派人", "被指派处理人"),
        ("assignment_note", "指派备注", "指派说明"),
        ("assignment_time", "指派时间", "指派时间"),
        ("processing_result", "处理结果", "处理结果描述"),
        ("completion_time", "完成时间", "处理完成时间"),
        ("created_at", "创建时间", "记录创建时间"),
        ("updated_at", "更新时间", "记录最后更新时间"),
    ]),
]

APIS = [
    ("users", "用户登录", "/api/auth/login", "POST", "Body: { username, password }", "LoginResponse { userId, username, realName, role, token }"),
    ("users", "用户注册", "/api/auth/register", "POST", "Body: { username, password, realName, role, phone }", '{ message: "注册成功" }'),
    ("users", "获取当前用户信息", "/api/auth/profile", "GET", "Header: Authorization", "User 对象"),
    ("users", "修改密码", "/api/auth/password", "PUT", "Header + Body: { oldPassword, newPassword }", "{ message }"),
    ("users", "更新个人信息", "/api/auth/profile", "PUT", "Header + Body: User", "{ message }"),
    ("users", "用户列表（分页）", "/api/users", "GET", "Query: page, pageSize, username?, role?, status?", "PageInfo<User>"),
    ("users", "创建用户", "/api/users", "POST", "Body: User", "User"),
    ("users", "更新用户", "/api/users/{id}", "PUT", "Path: id; Body: User", "User"),
    ("users", "删除用户", "/api/users/{id}", "DELETE", "Path: id", '{ message: "删除成功" }'),
    ("users", "用户详情", "/api/users/{id}", "GET", "Path: id", "User"),
    ("users", "重置密码", "/api/users/{id}/reset-password", "POST", "Path: id", '{ message: "密码已重置为默认密码: 123456" }'),
    ("users", "更新用户状态", "/api/users/{id}/status", "PUT", "Path: id; Query: status(1/0)", "{ message }"),
    ("reservoirs", "获取全部水库", "/api/reservoirs", "GET", "无", "List<Reservoir>"),
    ("reservoirs", "水库详情", "/api/reservoirs/{id}", "GET", "Path: id", "Reservoir"),
    ("reservoirs", "创建水库", "/api/reservoirs", "POST", "Body: Reservoir", "Reservoir"),
    ("reservoirs", "更新水库", "/api/reservoirs/{id}", "PUT", "Path: id; Body: Reservoir", "Reservoir"),
    ("reservoirs", "删除水库", "/api/reservoirs/{id}", "DELETE", "Path: id", "{ message }"),
    ("reservoirs", "检查水库名称是否存在", "/api/reservoirs/checkName", "GET", "Query: name", "boolean"),
    ("water_situation", "水情列表（分页）", "/api/waterSituation/list", "GET", "Query: page, pageSize, reservoirName?, startDate?, endDate? 及范围筛选", "PageInfo<WaterSituation>"),
    ("water_situation", "创建水情", "/api/waterSituation/create", "POST", "Body: WaterSituation", "WaterSituation"),
    ("water_situation", "更新水情", "/api/waterSituation/update", "POST", "Body: WaterSituation", "WaterSituation"),
    ("water_situation", "删除水情", "/api/waterSituation/delete/{id}", "DELETE", "Path: id", "空 200"),
    ("water_situation", "水情详情", "/api/waterSituation/detail/{id}", "GET", "Path: id", "WaterSituation"),
    ("water_situation", "Excel 导入", "/api/waterSituation/import", "POST", "Form: file", "List<WaterSituation>"),
    ("water_situation", "批量导入", "/api/waterSituation/batchImport", "POST", "Form: file", "ImportResult"),
    ("water_situation", "检查库名是否存在", "/api/waterSituation/checkReservoirName", "GET", "Query: reservoirName", "boolean"),
    ("water_situation", "导出水情数据", "/api/waterSituation/export", "GET", "Query: format?(xlsx/csv) + 筛选条件", "文件流"),
    ("water_situation", "下载导入模板", "/api/waterSituation/template/download", "GET", "无", ".xlsx 文件流"),
    ("section_monitor", "监测数据列表", "/api/sectionMonitor/list", "GET", "Query: page, pageSize, monitorPointName?, reservoirName? 及指标范围", "PageInfo<SectionMonitor>"),
    ("section_monitor", "创建监测记录", "/api/sectionMonitor/create", "POST", "Body: SectionMonitor", "SectionMonitor"),
    ("section_monitor", "更新监测记录", "/api/sectionMonitor/update", "POST", "Body: SectionMonitor", "SectionMonitor"),
    ("section_monitor", "删除监测记录", "/api/sectionMonitor/delete/{id}", "DELETE", "Path: id", '"删除成功"'),
    ("section_monitor", "监测详情", "/api/sectionMonitor/detail/{id}", "GET", "Path: id", "SectionMonitor"),
    ("section_monitor", "Excel 导入", "/api/sectionMonitor/import", "POST", "Form: file", "List<SectionMonitor>"),
    ("section_monitor", "批量导入", "/api/sectionMonitor/batchImport", "POST", "Form: file", "ImportResult<SectionMonitor>"),
    ("section_monitor", "检查监测点名称", "/api/sectionMonitor/checkMonitorPointName", "GET", "Query: monitorPointName, reservoirName", "boolean"),
    ("section_monitor", "导出监测数据", "/api/sectionMonitor/export", "POST", "Query: format?(xlsx/csv)", "文件流"),
    ("section_monitor", "下载导入模板", "/api/sectionMonitor/template", "GET", "无", ".xlsx 文件流"),
    ("patrol_records", "巡检签到", "/api/patrol/checkin", "POST", "Body: { lat, lng, address, inspector, time? }", "{ message, recordId }"),
    ("patrol_records", "巡检记录列表", "/api/patrol/records", "GET", "Query: page, pageSize, status?, inspector?, startDate?, endDate?", "PageInfo<PatrolRecord>"),
    ("patrol_records", "巡检记录详情", "/api/patrol/records/{id}", "GET", "Path: id", "PatrolRecord"),
    ("patrol_records", "问题上报", "/api/patrol/report", "POST", "Body: { reservoirName, description, photos, severity, reporterName, reporterRole, latitude, longitude }", "{ message, recordId }"),
    ("patrol_records", "任务派发", "/api/patrol/assign", "POST", "Body: { recordId, inspector, note }", "{ message }"),
    ("patrol_records", "更新巡检记录", "/api/patrol/records/{id}", "PUT", "Path: id; Body: PatrolRecord", "{ message }"),
    ("patrol_records", "更新巡检状态", "/api/patrol/records/{id}/status", "PUT", "Path: id; Query: status, situationDescription?", "{ message }"),
    ("patrol_records", "删除巡检记录", "/api/patrol/records/{id}", "DELETE", "Path: id", "{ message }"),
    ("warning_records", "预警列表", "/api/warnings", "GET", "Query: page, pageSize, level?, status?, startDate?, endDate?", "PageInfo<WarningRecord>"),
    ("warning_records", "预警详情", "/api/warnings/{id}", "GET", "Path: id", "WarningRecord"),
    ("warning_records", "更新预警", "/api/warnings/{id}", "PUT", "Path: id; Body: WarningRecord", "{ message }"),
    ("warning_records", "更新预警状态", "/api/warnings/{id}/status", "PUT", "Path: id; Query: status", "{ message }"),
    ("warning_records", "删除预警", "/api/warnings/{id}", "DELETE", "Path: id", "{ message }"),
    ("warning_records", "从监测数据生成预警", "/api/warnings/generate", "POST", "无", '{ message: "生成N条预警记录" }'),
    ("warning_thresholds", "获取预警阈值", "/api/warning-thresholds", "GET", "无", "WarningThreshold"),
    ("warning_thresholds", "更新预警阈值", "/api/warning-thresholds", "PUT", "Body: WarningThreshold", "{ message }"),
    ("roles", "角色列表", "/api/roles", "GET", "无", "List<Role>"),
    ("roles", "角色详情", "/api/roles/{id}", "GET", "Path: id", "Role"),
    ("roles", "创建角色", "/api/roles", "POST", "Body: Role", "Role"),
    ("roles", "更新角色", "/api/roles/{id}", "PUT", "Path: id; Body: Role", "Role"),
    ("roles", "删除角色", "/api/roles/{id}", "DELETE", "Path: id", "{ message }"),
    ("simulation_configs", "配置列表", "/api/simulation-configs", "GET", "无", "List<SimulationConfig>"),
    ("simulation_configs", "默认配置", "/api/simulation-configs/default", "GET", "无", "SimulationConfig"),
    ("simulation_configs", "创建配置", "/api/simulation-configs", "POST", "Body: SimulationConfig", "SimulationConfig"),
    ("simulation_configs", "更新配置", "/api/simulation-configs/{id}", "PUT", "Path: id; Body: SimulationConfig", "{ message }"),
    ("simulation_configs", "删除配置", "/api/simulation-configs/{id}", "DELETE", "Path: id", "{ message }"),
    ("simulation_configs", "设为默认", "/api/simulation-configs/{id}/set-default", "PUT", "Path: id", "{ message }"),
    ("layer_configs", "获取全部图层配置", "/api/layer-configs", "GET", "无", "List<LayerConfig>"),
    ("layer_configs", "批量更新图层配置", "/api/layer-configs", "PUT", "Body: List<LayerConfig>", "{ message }"),
    ("layer_configs", "获取图层类型列表", "/api/layer-configs/types", "GET", "无", "List<String>"),
    ("inspection_tasks", "待办任务", "/api/mobile/tasks/pending", "GET", "Query: inspectorId?, inspector?", "List<InspectionTask>"),
    ("inspection_tasks", "任务列表", "/api/mobile/tasks/list", "GET", "Query: page, pageSize, status?, assigneeId?, assigneeName?", "PageInfo<InspectionTask>"),
    ("inspection_tasks", "任务详情", "/api/mobile/tasks/{id}", "GET", "Path: id", "InspectionTask"),
    ("inspection_tasks", "接受任务", "/api/mobile/tasks/{id}/accept", "POST", "Query: assigneeId, assigneeName", "{ message }"),
    ("inspection_tasks", "提交反馈", "/api/mobile/tasks/{id}/feedback", "POST", "Body: { content, photos, inspector, inspectorUsername }", "{ message }"),
    ("inspection_tasks", "完成任务", "/api/mobile/tasks/{id}/complete", "POST", "Body: { processingResult, description? }", "{ message }"),
    ("inspection_tasks", "Web 任务列表", "/api/web/tasks", "GET", "Query: 分页 + 筛选", "PageInfo<InspectionTask>"),
    ("inspection_tasks", "Web 创建任务", "/api/web/tasks/create", "POST", "Body: CreateTaskRequest", "{ message, taskId }"),
    ("inspection_tasks", "Web 更新任务", "/api/web/tasks/{id}", "PUT", "Path: id; Body: InspectionTask", "{ message }"),
    ("inspection_tasks", "Web 删除任务", "/api/web/tasks/{id}", "DELETE", "Path: id", "{ message }"),
    ("issue_reports", "我的上报", "/api/mobile/reports/my", "GET", "Query: reporter", "List<IssueReport>"),
    ("issue_reports", "待处理上报", "/api/mobile/reports/pending", "GET", "无", "List<IssueReport>"),
    ("issue_reports", "上报详情", "/api/mobile/reports/{id}", "GET", "Path: id", "IssueReport"),
    ("issue_reports", "创建上报", "/api/mobile/reports", "POST", "Body: IssueReport", "{ message, reportId }"),
    ("issue_reports", "处理上报", "/api/mobile/reports/{id}/process", "POST", "Body: { processingResult }", "{ message }"),
    ("issue_reports", "Web 上报列表", "/api/web/reports", "GET", "Query: page, pageSize, status?, severity?, startDate?, endDate?", "PageInfo<IssueReport>"),
    ("issue_reports", "Web 创建上报", "/api/web/reports", "POST", "Body: IssueReport", "{ message, reportId }"),
    ("issue_reports", "转为任务", "/api/web/reports/{id}/to-task", "POST", "Body: { assigneeId, assigneeName, note }", "{ message, taskId }"),
    ("issue_reports", "标记解决", "/api/web/reports/{id}/resolve", "POST", "Body: { processingResult }", "{ message }"),
    ("issue_reports", "指派处理人", "/api/web/reports/{id}/assign", "POST", "Body: { assignedInspector, assignmentNote }", "{ message }"),
    ("issue_reports", "删除上报", "/api/web/reports/{id}", "DELETE", "Path: id", "{ message }"),
    ("多表", "统计概览", "/api/dashboard/stats", "GET", "无", "DashboardStats"),
    ("section_monitor", "水质趋势", "/api/dashboard/water-quality/trend", "GET", "Query: reservoirId?, startDate?, endDate?", "List<WaterQualityTrend>"),
    ("warning_records", "最新预警", "/api/dashboard/warnings/latest", "GET", "Query: limit(默认5)", "List<WarningRecord>"),
    ("patrol_records", "巡检统计", "/api/dashboard/patrol/stats", "GET", "Query: date?", "{ todayCheckins, pendingTasks }"),
    ("—", "地图要素 GeoJSON", "/api/map/features", "GET", "Query: type?", "{ type, data, message }"),
    ("reservoirs", "水库点位 GeoJSON", "/api/map/features/reservoirs", "GET", "无", "{ data: GeoJSON }"),
    ("section_monitor", "监测点 GeoJSON", "/api/map/features/monitor-points", "GET", "Query: reservoirName?", "{ data: GeoJSON }"),
    ("—", "健康检查", "/actuator/health", "GET", "无", "Spring Actuator 健康状态 JSON"),
    ("—", "健康测试", "/api/test/hello", "GET", "无", '"Hello from backend!"'),
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, headers, widths):
    ws.append(headers)
    for col, width in enumerate(widths, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28


def style_data_rows(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_tables_sheet(wb):
    ws = wb.active
    ws.title = "数据表"
    style_header(ws, ["表名(英文)", "表中文名", "属性英文名", "中文名", "描述"], [18, 18, 22, 16, 40])
    for table_en, table_cn, fields in TABLES:
        for i, (field_en, field_cn, desc) in enumerate(fields):
            ws.append([
                table_en if i == 0 else "",
                table_cn if i == 0 else "",
                field_en,
                field_cn,
                desc,
            ])
    style_data_rows(ws)
    ws.freeze_panes = "A2"


def build_apis_sheet(wb):
    ws = wb.create_sheet("后端接口")
    style_header(ws, ["表名(英文)", "接口名称(中文)", "请求路径", "请求方法", "请求参数", "返回值"], [18, 22, 32, 10, 42, 36])
    for row in APIS:
        ws.append(list(row))
    style_data_rows(ws)
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    build_tables_sheet(wb)
    build_apis_sheet(wb)
    wb.save(OUTPUT)
    print(f"已导出: {OUTPUT}")


if __name__ == "__main__":
    main()
