#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 seed 目录 Excel 导入 clarimire 数据库
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pymysql

BASE = Path(__file__).resolve().parent
WATER_FILE = BASE / '水情表(2).xlsx'
SECTION_FILE = BASE / '监测断面数据（19-24）（创）.xlsx'

RESERVOIR_META = [
    ('白河堡水库', '白河堡', 592.6, 599.13, 591.6, 9060),
    ('半城子水库', '半城子', 255.0, 255.0, None, None),
    ('北台上水库', '北台上', 85.0, 84.8, None, None),
    ('崇青水库', '崇青', 71.5, 71.5, None, None),
    ('大宁水库', '大宁', 48.0, 59.29, None, None),
    ('大水峪水库', '大水峪', 166.4, 168.9, None, None),
    ('官厅水库', '官厅', 476.0, 497.0, None, 12000),
    ('海子水库', '海子', 106.5, 108.5, None, None),
    ('怀柔水库', '怀柔', 58.0, 62.13, None, 1200),
    ('黄松峪水库', '黄松峪', 203.0, 203.0, None, None),
    ('密云水库', '密云', 152.0, 155.59, 151.8, 43100),
    ('沙厂水库', '沙厂', 165.5, 165.5, None, None),
    ('十三陵水库', '十三陵', 93.0, 91.81, None, 7450),
    ('桃峪口水库', '桃峪口', 67.7, 70.23, None, None),
    ('西峪水库', '西峪', 213.5, 213.5, None, None),
    ('遥桥峪水库', '遥桥峪', 464.0, 464.0, None, None),
    ('斋堂水库', '斋堂', 453.0, 461.56, None, None),
    ('珠窝水库', '珠窝', 348.4, 352.5, None, None),
]

NAME_MAP = {
    '白河堡': '白河堡水库', '半城子': '半城子水库', '北台上': '北台上水库',
    '崇青': '崇青水库', '大宁': '大宁水库', '大  宁': '大宁水库',
    '大水峪': '大水峪水库', '官厅': '官厅水库', '官  厅': '官厅水库',
    '海子': '海子水库', '怀柔': '怀柔水库', '怀  柔': '怀柔水库',
    '黄松峪': '黄松峪水库', '密云': '密云水库', '密  云': '密云水库',
    '沙厂': '沙厂水库', '十三陵': '十三陵水库', '桃峪口': '桃峪口水库',
    '西峪': '西峪水库', '遥桥峪': '遥桥峪水库', '斋堂': '斋堂水库',
    '珠窝': '珠窝水库'
}


def normalize_name(name):
    if not name:
        return name
    n = re.sub(r'\s+', '', str(name).strip())
    return NAME_MAP.get(n, n if n.endswith('水库') else n + '水库')


def parse_water_date(text):
    if not text:
        return None
    s = str(text).strip()
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)), 8, 0, 0)
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s[:19], fmt)
        except ValueError:
            pass
    return None


def to_decimal(val):
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def import_water(conn):
    wb = openpyxl.load_workbook(WATER_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        name = normalize_name(row[0])
        dt = parse_water_date(row[1])
        if not dt:
            continue
        rows.append((
            name, dt,
            to_decimal(row[2]), to_decimal(row[3]), to_decimal(row[4]),
            to_decimal(row[5]), to_decimal(row[6]), to_decimal(row[7]),
            to_decimal(row[8])
        ))
    wb.close()
    with conn.cursor() as cur:
        cur.execute('TRUNCATE TABLE water_situation')
        cur.executemany(
            '''INSERT INTO water_situation
            (reservoir_name, date, water_level, storage, avg_inflow, avg_outflow,
             yoy_increase, total_capacity, flood_level)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
            rows
        )
    conn.commit()
    return len(rows)


def import_section(conn):
    wb = openpyxl.load_workbook(SECTION_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows.append((
            str(row[0]).strip(),
            normalize_name(row[1]),
            int(row[2]), int(row[3]),
            to_decimal(row[4]),  # 氨氮 -> ammonia_nitrogen
            to_decimal(row[5]),
            to_decimal(row[6]),
            to_decimal(row[7]),
            to_decimal(row[8]),
            to_decimal(row[9]),
            to_decimal(row[10])
        ))
    wb.close()
    with conn.cursor() as cur:
        cur.execute('TRUNCATE TABLE section_monitor')
        cur.executemany(
            '''INSERT INTO section_monitor
            (monitor_point_name, reservoir_name, year, month,
             ammonia_nitrogen, potassium_permanganate, cod, flow, water_depth,
             total_nitrogen, total_phosphorus)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
            rows
        )
    conn.commit()
    return len(rows)


def import_reservoirs(conn):
    with conn.cursor() as cur:
        cur.execute('DELETE FROM water_reservoir')
        cur.executemany(
            '''INSERT INTO water_reservoir
            (reservoir_name, short_name, flood_level, max_level, avg_level, capacity, status)
            VALUES (%s,%s,%s,%s,%s,%s,1)''',
            RESERVOIR_META
        )
    conn.commit()
    return len(RESERVOIR_META)


def main():
    conn = pymysql.connect(
        host='localhost', user='root', password='root',
        database='clarimire', charset='utf8mb4'
    )
    try:
        r = import_reservoirs(conn)
        w = import_water(conn)
        s = import_section(conn)
        print(f'导入完成: 水库 {r} 座, 水情 {w} 条, 监测断面 {s} 条')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
